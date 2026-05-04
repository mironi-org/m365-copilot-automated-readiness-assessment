"""
Module for exporting recommendations to CSV and Excel
"""
import csv
import json
import os
from datetime import datetime
from pathlib import Path

def export_to_csv(recommendations, filename=None):
    """
    Export recommendations to CSV file
    
    Args:
        recommendations: List of recommendation dictionaries
        filename: Output filename (optional, generates timestamp-based name if not provided)
    
    Returns:
        str: Path to created CSV file
    """
    # Create Reports folder if it doesn't exist
    recommendations_dir = Path("Reports")
    recommendations_dir.mkdir(exist_ok=True)
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"m365_recommendations_{timestamp}.csv"
    
    if not filename.endswith('.csv'):
        filename += '.csv'
    
    # Build full path
    filepath = recommendations_dir / filename
    
    if not recommendations:
        print("No recommendations to export.")
        return None
    
    # Define CSV headers
    fieldnames = ["Service", "Feature", "Status", "Priority", "Observation", "Recommendation", "LinkText", "LinkUrl"]
    
    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(recommendations)
    
    return str(filepath)

def export_to_json(recommendations, filename=None):
    """
    Export recommendations to JSON file
    
    Args:
        recommendations: List of recommendation dictionaries
        filename: Output filename (optional, generates timestamp-based name if not provided)
    
    Returns:
        str: Path to created JSON file
    """
    # Create Reports folder if it doesn't exist
    recommendations_dir = Path("Reports")
    recommendations_dir.mkdir(exist_ok=True)
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"m365_recommendations_{timestamp}.json"
    
    if not filename.endswith('.json'):
        filename += '.json'
    
    # Build full path
    filepath = recommendations_dir / filename
    
    if not recommendations:
        print("No recommendations to export.")
        return None
    
    with open(filepath, 'w', encoding='utf-8') as jsonfile:
        json.dump(recommendations, jsonfile, indent=2, ensure_ascii=False)
    
    print(f"Recommendations exported to JSON: {filepath}")
    return str(filepath)

def export_to_excel(recommendations, filename=None):
    """
    Export recommendations to Excel file
    Requires openpyxl: pip install openpyxl
    
    Args:
        recommendations: List of recommendation dictionaries
        filename: Output filename (optional, generates timestamp-based name if not provided)
    
    Returns:
        str: Path to created Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Warning: openpyxl not installed. Install it with: pip install openpyxl")
        print("Falling back to CSV export...")
        return export_to_csv(recommendations, filename)
    
    # Create Reports folder if it doesn't exist
    recommendations_dir = Path("Reports")
    recommendations_dir.mkdir(exist_ok=True)
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"m365_recommendations_{timestamp}.xlsx"
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Build full path
    filepath = recommendations_dir / filename
    
    if not recommendations:
        print("No recommendations to export.")
        return None
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    
    # Define headers
    headers = ["Service", "Feature", "Status", "Priority", "Observation", "Recommendation", "Link Text", "Link URL"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Priority colors
    priority_colors = {
        "High": "FF6B6B",
        "Medium": "FFD93D",
        "Low": "95E1D3"
    }
    
    # Columns that need text wrapping (1-based: E=5 Observation, F=6 Recommendation)
    wrap_columns = {5, 6}

    # Add data rows
    for rec in recommendations:
        row = [
            rec.get("Service", ""),
            rec.get("Feature", ""),
            rec.get("Status", ""),
            rec.get("Priority", ""),
            rec.get("Observation", ""),
            rec.get("Recommendation", ""),
            rec.get("LinkText", ""),
            rec.get("LinkUrl", "")
        ]
        ws.append(row)

        row_num = ws.max_row
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

        # Color code priority
        priority = rec.get("Priority", "")
        if priority in priority_colors:
            priority_cell = ws.cell(row=row_num, column=4)
            priority_cell.fill = PatternFill(start_color=priority_colors[priority],
                                            end_color=priority_colors[priority],
                                            fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Service
    ws.column_dimensions['B'].width = 40  # Feature
    ws.column_dimensions['C'].width = 15  # Status
    ws.column_dimensions['D'].width = 12  # Priority
    ws.column_dimensions['E'].width = 50  # Observation
    ws.column_dimensions['F'].width = 50  # Recommendation
    ws.column_dimensions['G'].width = 30  # Link Text
    ws.column_dimensions['H'].width = 60  # Link URL
    
    # Save workbook
    wb.save(filepath)
    return str(filepath)

def print_recommendations_summary(recommendations, csv_path=None, excel_path=None):
    """
    Print a summary of recommendations grouped by service and priority
    
    Args:
        recommendations: List of recommendation dictionaries
        csv_path: Path to CSV export (optional)
        excel_path: Path to Excel export (optional)
    """
    if not recommendations:
        print("\nℹ️  No recommendations generated")
        return
    
    print("\n" + "="*80)
    print("RECOMMENDATIONS SUMMARY")
    print("="*80)
    
    # Group by priority
    high_priority = [r for r in recommendations if r.get("Priority") == "High"]
    medium_priority = [r for r in recommendations if r.get("Priority") == "Medium"]
    low_priority = [r for r in recommendations if r.get("Priority") == "Low"]
    
    print(f"\nTotal Recommendations: {len(recommendations)}")
    print(f"  🔴 High Priority:   {len(high_priority)}")
    print(f"  🟡 Medium Priority: {len(medium_priority)}")
    print(f"  🟢 Low Priority:    {len(low_priority)}")
    
    # Group by service
    services = {}
    for rec in recommendations:
        service = rec.get("Service", "Unknown")
        if service not in services:
            services[service] = []
        services[service].append(rec)
    
    print(f"\nRecommendations by Service:")
    for service, recs in sorted(services.items()):
        print(f"  • {service}: {len(recs)} recommendation(s)")
    
    if csv_path and excel_path:
        print(f"\nRecommendations exported to CSV: {csv_path}")
        print(f"Recommendations exported to Excel: {excel_path}")
    print()