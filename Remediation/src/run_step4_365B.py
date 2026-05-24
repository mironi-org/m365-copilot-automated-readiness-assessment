"""
Step 4 — Audience-Specific Business Communication Email Templates

Generates Email_Templates.txt with two versions:
  1. Standard (IT Managers, Project Leads, Technical Contacts)
  2. Executive (C-Level, VPs, Directors, Decision Makers)

Usage:
    python run_step4_365B.py <input_excel> <output_directory>
"""

import sys
import os
import re
import openpyxl

if len(sys.argv) < 3:
    print("ERROR: Both parameters are required.")
    print("  <input_excel>       - Path to the assessment Excel file (REQUIRED)")
    print("  <output_directory>  - Path to the output directory (REQUIRED)")
    print()
    print("Usage: python run_step4_365B.py <input_excel> <output_directory>")
    sys.exit(1)

EXCEL_PATH = sys.argv[1]
OUTPUT_DIR = sys.argv[2]

if not os.path.isfile(EXCEL_PATH):
    print(f"ERROR: Input file not found: {EXCEL_PATH}")
    sys.exit(1)

if not OUTPUT_DIR.strip():
    print("ERROR: Output directory cannot be empty.")
    sys.exit(1)


def load_data(excel_path):
    """Load assessment data and return (headers, rows) from the active sheet."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append([str(v or "").strip() for v in row])
    wb.close()
    return headers, rows


def find_col(headers, keywords):
    for i, h in enumerate(headers):
        for kw in keywords:
            if kw.lower() in h.lower():
                return i
    return None


def generate_email_templates(excel_path, output_dir):
    """Generate audience-specific email templates based on assessment data."""
    os.makedirs(output_dir, exist_ok=True)

    # Load data
    headers, rows = load_data(excel_path)
    col_service = find_col(headers, ["Service"])
    col_feature = find_col(headers, ["Feature"])
    col_status = find_col(headers, ["Status"])
    col_priority = find_col(headers, ["Priority"])
    col_observation = find_col(headers, ["Observation"])

    # Filter non-success items
    non_success = [r for r in rows if col_status is not None and r[col_status] != "Success"]

    # Count by priority
    priority_counts = {}
    for r in non_success:
        p = r[col_priority] if col_priority is not None else "Unknown"
        priority_counts[p] = priority_counts.get(p, 0) + 1

    total_items = len(non_success)
    high_count = priority_counts.get("High", 0) + priority_counts.get("Critical", 0)
    medium_count = priority_counts.get("Medium", 0)
    low_count = priority_counts.get("Low", 0)

    # Count by service
    service_counts = {}
    for r in non_success:
        s = r[col_service] if col_service is not None else "Unknown"
        service_counts[s] = service_counts.get(s, 0) + 1

    # Get top high-priority items for executive summary
    high_items = [r for r in non_success if col_priority is not None and r[col_priority] in ("High", "Critical")]

    # Build high-priority bullet points
    high_bullets = ""
    for i, r in enumerate(high_items[:5], 1):
        feature = r[col_feature] if col_feature is not None else "Unknown"
        obs = r[col_observation] if col_observation is not None else ""
        high_bullets += f"   {i}. {feature} — {obs[:120]}\n"

    # Service breakdown text
    svc_text = ", ".join(f"{svc} ({cnt} items)" for svc, cnt in sorted(service_counts.items(), key=lambda x: -x[1]))

    email_content = f"""{'='*80}
EMAIL VERSION 1: STANDARD
(For IT Managers, Project Leads, Technical Contacts)
{'='*80}

Subject: Microsoft 365 Copilot Readiness - Action Tracker & Next Steps

Hi [Customer Name],

Following our comprehensive Microsoft 365 Copilot readiness assessment, I'm pleased to deliver your complete action tracker and implementation guidance package.

📊 Assessment Summary:
• Total remediation items: {total_items}
• High/Critical priority: {high_count} items (action within 30 days)
• Medium priority: {medium_count} items (30-90 day window)
• Low priority: {low_count} items (90+ days)
• Services affected: {svc_text}

📊 What's Included:
• All Items Tracker — Complete view of all {total_items} remediation items with tracking columns, assignment management, and progress monitoring
• Critical & High Priority — Your top {high_count} urgent items requiring attention within 30 days
• Summary Dashboard — Real-time metrics, priority distribution chart, and status breakdown for executive reporting
• By Service — Items grouped by service for team-based assignment
• Instructions — Complete user guide with column descriptions, workflow process, and priority timelines

✓ Key Features:
• Color-coded priorities — Red (Critical), Orange (High), Yellow (Medium), Green (Low) for instant visual scanning
• Dropdown status tracking — Not Started → In Progress → Completed → Blocked
• Auto-generated success criteria — Feature-specific validation checkpoints for each item
• Progress % column — Validated 0-100 numeric entry with color-coded completion bands
• Direct hyperlinks — Clickable links to Microsoft Learn documentation for every item

🚀 Quick Start (5 Steps):
1. Open the 'Critical & High Priority' sheet — Focus here first
2. Assign owners to each item using the 'Assigned To' column
3. Set target dates based on priority timelines: High = 30 days, Medium = 90 days
4. Begin with the highest-priority items immediately
5. Schedule weekly standup using the 'Summary Dashboard' as your progress agenda

⚠️ Priority Timeline:
• High/Critical ({high_count} items): 7-30 days — Immediate security and infrastructure gaps
• Medium ({medium_count} items): 30-90 days — Governance and optimization improvements
• Low ({low_count} items): 90+ days — Long-term enhancements

📌 Top Priority Items:
{high_bullets}
💼 Next Actions:
• Schedule kickoff meeting with identity team and productivity team by [Date + 3 days]
• Assign owners for all High priority items by [Date + 5 days]
• Begin immediate remediation of critical security gaps

📎 Attached Files:
• Copilot_Readiness_Tracker.xlsx — Project management workbook (5 sheets, conditional formatting, charts)
• Executive_Summary.docx — Assessment overview document for leadership
• Copilot_Readiness_Action_Plan.md — Detailed step-by-step implementation guide with success criteria
• High_Priority_Action_Plan.docx — Focused document for items needing immediate attention
• Medium_Priority_Action_Plan.docx — Comprehensive guide for medium-term governance items
• Low_Priority_Action_Plan.docx — Long-term enhancement roadmap
• Visualization PNGs — Status distribution, priority chart, service breakdown

I'm available for a walkthrough call this week to review the tracker together, discuss team assignments, and answer any implementation questions. Would [Day] at [Time] work for a 30-minute session?

Best regards,
[Your Name]
[Your Title]
[Contact Information]

{'='*80}
EMAIL VERSION 2: EXECUTIVE
(For C-Level, VPs, Directors, Decision Makers)
{'='*80}

Subject: Copilot Readiness Assessment - Action Tracker Delivered

Hi [Executive Name],

Your Microsoft 365 Copilot readiness assessment is complete. We've identified {total_items} remediation items with a structured execution tracker for your team.

🎯 Key Findings:
• {high_count} High/Critical priority items requiring action within 30 days
• {medium_count} Medium priority governance and optimization improvements (30-90 day window)
• {low_count} Low priority long-term enhancements (90+ days)
• Services: {svc_text}

📌 Top Priority Items:
{high_bullets}
📊 Deliverables Attached:
• Excel Tracker — Enterprise project management tool with color-coded priorities, dropdown tracking, charts, and team assignment views (5 sheets)
• Executive Summary — Assessment overview with strategic recommendations and service analysis
• Implementation Guide — Detailed step-by-step remediation instructions with success criteria
• Priority Word Documents — Self-contained action plans segmented by priority level (High/Medium/Low)
• Visualizations — Status distribution, priority breakdown, and service analysis charts

⚡ Immediate Actions Required:
{high_bullets}
Your team should use the 'Critical & High Priority' sheet in the Excel tracker as their daily focus board. We recommend weekly 15-minute progress reviews using the built-in Summary Dashboard until all High items are resolved.

I'd welcome a brief 15-minute alignment call this week to ensure your teams have everything they need to begin execution. What does your availability look like on [Day] or [Day]?

Best regards,
[Your Name]
[Your Title]

📎 Attachments: Tracker (Excel) | Executive Summary (Word) | Implementation Guide (Markdown) | Priority Docs (3x Word) | Visualizations (PNGs)
"""

    output_path = os.path.join(output_dir, "Email_Templates.txt")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(email_content)

    print(f"  ✓ Email_Templates.txt created (2 versions, {len(email_content):,} chars)")
    return output_path, len(email_content)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"  STEP 4: Audience-Specific Email Templates")
    print(f"{'='*60}")
    print(f"  Input:  {EXCEL_PATH}")
    print(f"  Output: {OUTPUT_DIR}")
    print(f"{'='*60}\n")

    path, size = generate_email_templates(EXCEL_PATH, OUTPUT_DIR)

    print(f"\n{'='*60}")
    print(f"  Step 4 Complete")
    print(f"  Output: {path}")
    print(f"  Size:   {size:,} characters")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
