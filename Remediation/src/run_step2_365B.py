"""
Step 2: Detailed Row-by-Row Implementation Guide
From: M365_Copilot_Assessment_Insights_Generator.md (Option B)
Input: <input_excel> (passed as parameter)
Output: <output_directory>/Copilot_Readiness_Action_Plan.md
"""

import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ─── CONFIG ───────────────────────────────────────────────────────────────────
if len(sys.argv) < 3:
    print("ERROR: Both parameters are required.")
    print("  <input_excel>       - Path to the assessment Excel file (REQUIRED)")
    print("  <output_directory>  - Path to the output directory (REQUIRED)")
    print()
    print("Usage: python run_step2_365B.py <input_excel> <output_directory>")
    sys.exit(1)

EXCEL_PATH = sys.argv[1]
OUTPUT_DIR = sys.argv[2]

if not os.path.isfile(EXCEL_PATH):
    print(f"ERROR: Input file not found: {EXCEL_PATH}")
    sys.exit(1)

if not OUTPUT_DIR.strip():
    print("ERROR: Output directory cannot be empty.")
    sys.exit(1)


# ─── RISK ASSESSMENT LOGIC (aligned with reference implementation) ────────────
def assess_risk(feature, observation, status):
    """
    Generate risk assessment columns based on Status + Observation signals.
    Logic derived from original reference implementation:
    - Primary axis: Status value (how did the assessment tool collect data?)
    - Secondary axis: Observation content (what type of data was obtained?)
    - FN Risk varies by feature family (SharePoint/Teams/Intune/generic)
    Returns dict with: telemetry, fp_risk, fn_risk, confidence, impact
    """
    f = (feature or "").lower()
    o = (observation or "").lower()
    s = (status or "").lower()

    # ── Determine FN Risk by feature family ──
    if "sharepoint" in f or "onedrive" in f or "site" in f or "content" in f:
        fn_risk = "Medium"
    elif "teams" in f or "meeting" in f or "collaboration" in f:
        fn_risk = "Medium"
    elif "intune" in f or "device management" in f or "compliance polic" in f:
        fn_risk = "Medium"
    else:
        fn_risk = "Medium"

    # ── Determine impact by feature type ──
    if "conditional access" in f or "mfa" in f or "multi-factor" in f:
        impact = "High — identity security gap affects Copilot access controls"
    elif "sharepoint" in f or "onedrive" in f or "content" in f:
        impact = "High — insufficient content limits Copilot knowledge base"
    elif "defender" in f or "endpoint" in f or "xdr" in f:
        impact = "High — unmanaged devices create data protection blind spots"
    elif "sensitivity label" in f or "dlp" in f or "information protection" in f:
        impact = "High — missing data protection policies expose sensitive data via Copilot"
    else:
        impact = "Medium — may affect Copilot readiness posture"

    # ── Rule 1: Access Denied — couldn't collect data at all ──
    if "access denied" in s:
        return {
            "telemetry": "Unknown",
            "fp_risk": "Medium",
            "fn_risk": "Medium",
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Rule 2: Warning (especially Defender threats) — partial telemetry ──
    if "warning" in s:
        return {
            "telemetry": "Telemetry Partial",
            "fp_risk": "High",
            "fn_risk": "High",
            "confidence": "Low",
            "impact": impact,
        }

    # ── Rule 3: Critical status — derived posture score ──
    if "critical" in s:
        return {
            "telemetry": "Derived Posture",
            "fp_risk": "Medium",
            "fn_risk": "High",
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Rule 4: PendingActivation — service exists but API blocked ──
    if "pendingactivation" in s:
        return {
            "telemetry": "Permission Blocked",
            "fp_risk": "High",
            "fn_risk": "Medium",
            "confidence": "Low",
            "impact": impact,
        }

    # ── Rule 5: PendingInput — needs manual verification ──
    if "pendinginput" in s:
        return {
            "telemetry": "Manual Verification Required",
            "fp_risk": "High",
            "fn_risk": "Medium",
            "confidence": "Low",
            "impact": impact,
        }

    # ── Rule 6: Missing Prerequisite / Permission Required ──
    if "missing prerequisite" in s or "permission required" in s:
        return {
            "telemetry": "Unknown",
            "fp_risk": "Medium",
            "fn_risk": "Medium",
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Rule 7: Disabled — config absence confirmed via API ──
    if "disabled" in s:
        return {
            "telemetry": "Configuration Gap",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Rule 8: Action Required — config gap confirmed ──
    if "action required" in s:
        return {
            "telemetry": "Configuration Gap",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Rule 9: Insight with zero counts — inferred from metrics ──
    if "insight" in s:
        # Check if observation has zero-count patterns
        if " 0 " in o or "0 active" in o or "0 teams" in o or "0 sharepoint" in o or "0 sites" in o:
            return {
                "telemetry": "Index Inferred",
                "fp_risk": "Medium",
                "fn_risk": fn_risk,
                "confidence": "Medium",
                "impact": impact,
            }
        return {
            "telemetry": "Index Inferred",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
            "impact": impact,
        }

    # ── Default — standard Graph API read ──
    return {
        "telemetry": "Index Inferred",
        "fp_risk": "Medium",
        "fn_risk": fn_risk,
        "confidence": "Medium",
        "impact": impact,
    }


# ─── COPILOT READINESS CONTEXT ───────────────────────────────────────────────
def get_copilot_context(feature, status):
    """Generate a context paragraph explaining why this matters for Copilot readiness."""
    f = (feature or "").lower()

    if "conditional access" in f or "ca polic" in f:
        return "Conditional Access policies are foundational for Copilot deployment. Without proper CA policies, organizations cannot enforce secure authentication requirements for Copilot access, leaving data exposed to unauthorized users and unmanaged devices."
    elif "mfa" in f or "multi-factor" in f or "multifactor" in f:
        return "MFA is a prerequisite for secure Copilot adoption. Copilot processes sensitive organizational data, and without MFA, compromised accounts could leverage Copilot to access and exfiltrate information across the entire M365 ecosystem."
    elif "sharepoint" in f or "onedrive" in f or "content" in f or "site" in f:
        return "SharePoint content is the primary knowledge base for Microsoft 365 Copilot. Without properly structured and indexed content in SharePoint, Copilot lacks the organizational knowledge to generate relevant, accurate responses for users."
    elif "defender" in f or "endpoint" in f or "xdr" in f or "device" in f:
        return "Device security is critical for Copilot deployments. Unmanaged or unprotected devices accessing Copilot create data exfiltration risks, as Copilot can surface sensitive data that may be compromised on insecure endpoints."
    elif "sensitivity label" in f or "dlp" in f or "information protection" in f:
        return "Data protection policies directly govern what Copilot can access and surface. Without proper sensitivity labels and DLP policies, Copilot may inadvertently expose classified information or generate responses containing restricted data."
    elif "purview" in f or "compliance" in f or "audit" in f:
        return "Compliance and audit capabilities are essential for Copilot governance. Organizations must demonstrate that Copilot interactions comply with regulatory requirements and that AI-generated content is auditable."
    elif "teams" in f or "meeting" in f or "collaboration" in f:
        return "Teams integration is a key Copilot surface area. Meeting summaries, chat assistance, and collaboration features depend on properly configured Teams settings and adequate usage patterns."
    elif "intune" in f or "device management" in f or "compliance polic" in f:
        return "Device compliance policies ensure that only secure, managed devices can access Copilot. Without Intune enforcement, sensitive data processed by Copilot may be accessible from non-compliant or compromised devices."
    elif "oauth" in f or "app" in f or "consent" in f or "application" in f:
        return "Application governance is critical for Copilot security. Over-privileged or malicious apps can access the same data as Copilot, potentially creating unauthorized data flows or shadow AI that bypasses organizational controls."
    elif "access review" in f or "governance" in f:
        return "Access reviews ensure that Copilot users maintain appropriate permissions. Without regular reviews, users may retain access to sensitive data through Copilot that they no longer need, violating the principle of least privilege."
    elif "password" in f or "passwordless" in f:
        return "Modern authentication strengthens the security posture for Copilot access. Passwordless methods reduce phishing risk, which is especially important when Copilot can surface sensitive organizational data."
    else:
        return "This configuration contributes to the overall security and readiness posture required for Microsoft 365 Copilot deployment. Addressing this gap helps ensure Copilot operates within a properly secured environment."


# ─── STEP-BY-STEP IMPLEMENTATION ─────────────────────────────────────────────
def get_steps_by_feature(feature, service):
    """Generate context-aware implementation steps based on feature type."""
    f = (feature or "").lower()

    if "conditional access" in f or "ca polic" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Protection → Conditional Access → Policies",
            "Click **+ New policy** and provide a descriptive name following naming convention (e.g., 'CA001 - Block Legacy Auth')",
            "Under **Assignments → Users**: Select target users/groups (start with pilot group, then expand to All users)",
            "Under **Target resources → Cloud apps**: Select 'Office 365' or specific apps as needed",
            "Under **Conditions**: Configure sign-in risk, device platforms, locations, and client apps as appropriate",
            "Under **Grant**: Select required controls (e.g., 'Require multifactor authentication', 'Require compliant device')",
            "Set policy to **Report-only** mode first — this allows monitoring without enforcement",
            "Monitor for 7-14 days: Review sign-in logs at Entra admin center → Monitoring → Sign-in logs (filter by CA policy)",
            "Validate no unexpected blocks or legitimate user lockouts in the pilot group",
            "Switch policy from Report-only to **On** after successful validation period",
            "Document the policy settings, scope, exclusions, and exceptions in your change management system",
            "Schedule a 30-day review to assess policy effectiveness, false positive rate, and user impact",
        ]
    elif "mfa" in f or "multi-factor" in f or "multifactor" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Protection → Authentication methods → Policies",
            "Enable required MFA methods: Microsoft Authenticator (recommended primary), FIDO2 keys, SMS (backup only)",
            "Configure **Microsoft Authenticator** settings: Enable number matching, additional context, and passwordless mode",
            "Navigate to Protection → Conditional Access and create/update policy requiring MFA for Office 365 access",
            "Prepare user communication: Announce enrollment timeline with clear deadlines and registration guides",
            "Direct users to [https://aka.ms/mysecurityinfo](https://aka.ms/mysecurityinfo) to register authentication methods",
            "Ensure each user registers at least 2 methods (primary + backup) for account recovery resilience",
            "Prepare help desk with enrollment support procedures, common issues, and escalation paths",
            "Monitor enrollment progress: Entra admin center → Protection → Authentication methods → User registration details",
            "After enrollment deadline, enforce via Conditional Access (block non-compliant sign-ins)",
            "Track compliance: Target 100% enrollment within 30 days of campaign start",
        ]
    elif "sharepoint" in f or "content" in f or "site" in f:
        return [
            "Assess current SharePoint environment: Navigate to [SharePoint admin center](https://admin.microsoft.com/sharepoint) → Active sites",
            "Identify content sources for migration: file shares, network drives, legacy platforms, local documents",
            "Plan site architecture: Create departmental/project sites with clear naming conventions and proper structure",
            "Create sites: SharePoint admin center → Sites → Active sites → + Create (choose Team site or Communication site)",
            "Configure site settings: Permissions model (private/public), sharing policies, storage quotas",
            "Use SharePoint Migration Tool (SPMT) or Migration Manager for bulk content migration from file shares",
            "Apply metadata, content types, and managed columns to migrated content for Copilot discoverability",
            "Configure governance: Apply sensitivity labels, retention policies, and DLP rules to sites/libraries",
            "Ensure minimum 100+ documents are properly indexed — verify via SharePoint search works correctly",
            "Validate Copilot can search and reference migrated content: Test with Copilot prompts about uploaded docs",
            "Create training materials and conduct user adoption sessions for content owners and editors",
            "Set up usage metrics tracking: SharePoint admin center → Reports → Site usage",
        ]
    elif "defender" in f or "xdr" in f or "device" in f or "endpoint" in f:
        return [
            "Navigate to [Microsoft Defender portal](https://security.microsoft.com) → Settings → Endpoints → Onboarding",
            "Select deployment method for your environment: Intune (recommended), GPO, local script, or SCCM",
            "For Intune: Create EDR configuration profile under Endpoint security → Endpoint detection and response",
            "For GPO: Download onboarding script package and deploy via Group Policy to target OUs",
            "Start with pilot group (10-20 devices across different departments) to validate onboarding",
            "Verify onboarded devices: Defender portal → Assets → Devices → Device inventory (check status = Onboarded)",
            "Configure security policies: Defender portal → Settings → Endpoints → Configuration management",
            "Enable Automated Investigation and Response (AIR): Settings → Endpoints → Advanced features → toggle On",
            "Create custom detection rules for organization-specific threat indicators and behaviors",
            "Configure alert notifications: Settings → Email notifications → Add recipient groups for Critical/High alerts",
            "Target: 50%+ device onboarding within first 30 days, 90%+ within 90 days",
            "Train SOC team on Defender portal investigation workflows, incident management, and response playbooks",
        ]
    elif "sensitivity label" in f or "dlp" in f or "information protection" in f or "data loss" in f:
        return [
            "Navigate to [Microsoft Purview compliance portal](https://compliance.microsoft.com) → Information protection → Labels",
            "Design label taxonomy: Define 3-5 sensitivity levels (e.g., Public, Internal, Confidential, Highly Confidential)",
            "Create sensitivity labels with appropriate protection settings (encryption, watermarks, access restrictions)",
            "Publish labels via label policies: Target all users or specific groups for initial rollout",
            "Configure auto-labeling policies: Identify sensitive content types (SSN, credit card, etc.) for automatic classification",
            "Create DLP policies: Compliance portal → Data loss prevention → Policies → + Create policy",
            "Set DLP rules to detect and protect sensitive information in Exchange, SharePoint, OneDrive, and Teams",
            "Deploy in test mode first: Run DLP policies in simulation for 7 days to assess impact",
            "Review DLP reports: Activity explorer → Filter by policy matches and false positives",
            "Enable enforcement after validation: Switch policies from test to enforce mode",
            "Communicate to users: Explain labeling expectations, provide quick-reference guides",
            "Monitor adoption: Compliance portal → Reports → Label analytics",
        ]
    elif "oauth" in f or "app" in f or "consent" in f or "application" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Applications → Enterprise applications",
            "Review all registered applications: Sort by 'Last sign-in' to identify active vs. stale apps",
            "Identify over-privileged apps: Check granted permissions vs. actually needed (focus on Mail.ReadWrite, Files.ReadWrite.All)",
            "For high-risk apps (excessive permissions, unknown publishers, no recent sign-in): Plan revocation",
            "Revoke consent: Select app → Permissions → Review permissions → Revoke admin consent",
            "Configure admin consent workflow: Entra → Applications → Consent and permissions → Admin consent settings → Enable",
            "Set consent policy: Block user consent for unverified publishers; require admin approval",
            "Create app governance alerting: Monitor for new apps requesting high-privilege permissions",
            "Set up monitoring: Entra → Monitoring → Audit logs (filter by Application consent category)",
            "Schedule quarterly app review cadence with designated reviewers from Security and IT teams",
            "Document approved applications registry: App name, owner, required permissions, review date",
        ]
    elif "password" in f or "passwordless" in f or "authentication method" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Protection → Authentication methods → Policies",
            "Enable passwordless methods: Microsoft Authenticator (passwordless mode), FIDO2 security keys, Windows Hello for Business",
            "Configure Microsoft Authenticator: Enable 'Allow passwordless sign-in' toggle",
            "For FIDO2: Add allowed security key models under FIDO2 security key → Key Restriction Policy",
            "Select pilot group (5-10 IT/security staff) to validate passwordless experience across all scenarios",
            "Direct pilot users to [https://aka.ms/mysecurityinfo](https://aka.ms/mysecurityinfo) to register passwordless methods",
            "Collect pilot feedback after 2 weeks: Usability, edge cases (VPN, legacy apps), and issues",
            "Create user training materials: Step-by-step registration guide, supported devices and scenarios",
            "Expand to larger group (50+ users) after successful pilot validation",
            "Monitor adoption: Authentication methods → User registration details → Filter by passwordless methods",
            "Track metrics: Passwordless sign-in success rate, user satisfaction, help desk ticket reduction",
        ]
    elif "access review" in f or "governance" in f or "review" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Identity Governance → Access reviews",
            "Click **+ New access review** and configure the review scope",
            "Select scope: Choose groups, applications, or Entra ID roles to review",
            "Set reviewers: Assign group owners, resource managers, or self-review as appropriate",
            "Configure recurrence: Quarterly (recommended for sensitive resources) or semi-annually",
            "Set review duration: 7-14 days for reviewers to complete their decisions",
            "Enable auto-apply: Configure auto-removal for denied or non-responded items after review closes",
            "Configure notifications: Email reminders at start, mid-point, and 2 days before deadline",
            "Run first review cycle: Monitor completion rate and reviewer participation",
            "After first cycle, review results: Verify decisions are appropriate, adjust scope/reviewers as needed",
            "Document access review program: Policies, schedules, escalation procedures",
        ]
    elif "cross-tenant" in f or "external" in f or "b2b" in f or "guest" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → External Identities → Cross-tenant access settings",
            "Review current default settings: Understand inbound/outbound access baseline",
            "Define organizational trust policies: Which partners get trusted MFA, compliant device claims",
            "Add organization-specific settings: Click + Add organization for key partners",
            "Configure inbound access: Define which external users can access your resources",
            "Configure outbound access: Define which resources your users can access externally",
            "Enable MFA trust for trusted partners: Inbound → Trust settings → Trust MFA from this organization",
            "Configure Conditional Access for external users: Create CA policy targeting 'All guest and external users'",
            "Test with partner organization: Verify access works correctly for intended scenarios",
            "Review B2B collaboration settings: External Identities → External collaboration settings",
            "Monitor external access: Entra → Monitoring → Sign-in logs → Filter by Cross-tenant access type",
        ]
    elif "intune" in f or "device management" in f or "compliance polic" in f or "enrollment" in f:
        return [
            "Navigate to [Microsoft Intune admin center](https://intune.microsoft.com) → Devices → Enrollment",
            "Configure enrollment restrictions: Define which device types and platforms are allowed",
            "Create compliance policies: Devices → Compliance → + Create policy (per platform: Windows, iOS, Android)",
            "Define compliance requirements: OS version, encryption, password complexity, threat level",
            "Create device configuration profiles: Devices → Configuration → + Create profile",
            "Configure app protection policies: Apps → App protection policies → + Create policy",
            "Link compliance to Conditional Access: Create CA policy requiring 'Require device to be marked as compliant'",
            "Deploy to pilot group first: Assign policies to a test group (10-20 devices) for 7 days",
            "Monitor compliance status: Devices → Monitor → Device compliance → Filter by non-compliant",
            "Address non-compliant devices: Contact device owners, provide remediation guidance",
            "Expand enrollment to all managed devices after successful pilot",
            "Set up regular compliance reporting: Weekly review of non-compliant device count",
        ]
    elif "group" in f or "licens" in f:
        return [
            "Navigate to [Microsoft Entra admin center](https://entra.microsoft.com) → Groups → All groups",
            "Identify or create target groups for license assignment (e.g., 'M365-E5-Users', 'Copilot-Licensed')",
            "Navigate to Entra → Billing → Licenses → Select the license product to assign",
            "Click Assign → Select group(s) → Choose specific service plans to enable/disable",
            "Verify license assignment: Groups → Select group → Licenses → Check processing status",
            "Handle assignment errors: Check for conflicts (e.g., user already has direct assignment)",
            "Remove direct user assignments that overlap with group-based assignments",
            "Monitor license utilization: Billing → Licenses → Usage reports",
            "Document group-license mapping: Create reference sheet of which groups get which licenses",
            "Set up dynamic groups for automatic license assignment based on user attributes",
        ]
    else:
        # Generic fallback
        return [
            "Review the current configuration state in the relevant admin portal",
            "Access the appropriate admin center: [Entra](https://entra.microsoft.com), [M365](https://admin.microsoft.com), [Security](https://security.microsoft.com), or [Compliance](https://compliance.microsoft.com)",
            "Consult the official Microsoft documentation linked in the Reference section below",
            "Plan implementation approach: Define scope, responsible parties, and target timeline",
            "Configure changes in a test/pilot scope first (limited user group or report-only mode)",
            "Validate changes produce expected results without adverse impact on users",
            "Monitor for 7-14 days post-implementation for any issues or regressions",
            "Document all changes made, including configuration details and rollback procedures",
            "Schedule follow-up review to confirm sustained compliance in next assessment cycle",
        ]


# ─── SUCCESS CRITERIA ─────────────────────────────────────────────────────────
def get_success_criteria(feature, service):
    """Generate feature-appropriate success criteria checkboxes."""
    f = (feature or "").lower()

    if "conditional access" in f or "ca polic" in f:
        return [
            "CA policy created and enabled (not in Report-only mode)",
            "Policy targets appropriate user scope (All users or defined groups)",
            "Office 365 cloud app included in target resources",
            "Required grant controls enforced (MFA, compliant device, etc.)",
            "Policy tested in Report-only for 7+ days with zero unexpected blocks",
            "Sign-in logs confirm policy is evaluating and enforcing correctly",
            "Break-glass/emergency accounts properly excluded from policy",
        ]
    elif "mfa" in f or "multi-factor" in f or "multifactor" in f:
        return [
            "100% of targeted users enrolled in MFA",
            "Each user has primary + backup authentication method registered",
            "Microsoft Authenticator configured as recommended primary method",
            "Conditional Access policy enforcing MFA is enabled",
            "Communication sent to all users with clear registration instructions",
            "Help desk prepared with enrollment support procedures",
            "Zero lockout incidents lasting >4 hours post-enforcement",
        ]
    elif "sharepoint" in f or "content" in f or "site" in f:
        return [
            "3-5 SharePoint sites created with proper structure and permissions",
            "Content migrated from legacy sources with intact metadata",
            "100+ documents indexed and searchable via SharePoint search",
            "Copilot can search and reference migrated content accurately",
            "Sensitivity labels and governance policies applied",
            "User training completed for content owners and editors",
            "Usage metrics showing active engagement post-migration",
        ]
    elif "defender" in f or "xdr" in f or "device" in f or "endpoint" in f:
        return [
            "50%+ of devices onboarded to Defender for Endpoint",
            "Device inventory visible and accurate in Defender portal",
            "Security policies (EDR, ASR) applied to onboarded devices",
            "Automated Investigation and Response (AIR) enabled",
            "Custom detection rules created for org-specific threats",
            "Alert notifications configured for security team",
            "SOC team trained on investigation and response workflows",
        ]
    elif "sensitivity label" in f or "dlp" in f or "information protection" in f:
        return [
            "Sensitivity label taxonomy defined and published",
            "Labels available to all targeted users in Office apps",
            "Auto-labeling policies configured for sensitive content types",
            "DLP policies created and enforcing across M365 workloads",
            "Test/simulation mode validated with acceptable false positive rate",
            "User communication and training materials distributed",
            "Label usage metrics tracked via compliance portal reports",
        ]
    elif "oauth" in f or "app" in f or "consent" in f or "application" in f:
        return [
            "All enterprise applications reviewed and risk-categorized",
            "Over-privileged or unused applications revoked",
            "Admin consent workflow enabled with designated approvers",
            "User consent restricted to verified publishers only",
            "App governance monitoring and alerting active",
            "Quarterly review cadence established with assigned owners",
        ]
    elif "password" in f or "passwordless" in f or "authentication method" in f:
        return [
            "Passwordless authentication methods enabled in policy",
            "Pilot group successfully using passwordless sign-in",
            "User feedback collected and issues addressed",
            "Training materials created and distributed to broader audience",
            "Adoption metrics tracked and trending upward",
        ]
    elif "access review" in f or "governance" in f or "review" in f:
        return [
            "Access review created with appropriate scope (groups/apps/roles)",
            "Quarterly recurrence configured and active",
            "Reviewers assigned, notified, and participating",
            "First review cycle completed successfully",
            "Auto-apply actions configured for denied/non-responded items",
            "Results documented and remediation actions tracked",
        ]
    elif "cross-tenant" in f or "external" in f or "b2b" in f or "guest" in f:
        return [
            "Cross-tenant access settings reviewed and configured per partner",
            "MFA trust enabled for verified partner organizations",
            "Conditional Access policy for guest users active",
            "B2B collaboration settings aligned with security policy",
            "External access monitoring in place via sign-in logs",
        ]
    elif "intune" in f or "device management" in f or "compliance polic" in f:
        return [
            "Device enrollment configured for all supported platforms",
            "Compliance policies created and assigned to device groups",
            "Conditional Access linked to device compliance requirement",
            "Non-compliant devices identified and remediation tracked",
            "Pilot validated; policies expanded to full scope",
            "Regular compliance reporting cadence established",
        ]
    elif "group" in f or "licens" in f:
        return [
            "Target groups created/identified for license assignment",
            "Licenses assigned to groups (not individual users)",
            "Assignment processing completed without errors",
            "Direct user assignments removed where overlapping",
            "License utilization monitored and documented",
        ]
    else:
        return [
            "Configuration completed as per recommendation",
            "Changes validated in the admin portal — correct settings confirmed",
            "No adverse impact on users or services observed during monitoring period",
            "Documentation updated with new configuration and rollback procedures",
            "Next assessment expected to show improved/resolved status",
        ]


# ─── MAIN GENERATION ─────────────────────────────────────────────────────────
def priority_sort_key(priority):
    order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    return order.get((priority or "").lower(), 99)


def generate_action_plan(excel_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    col_map = {h.lower(): i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and any(cell is not None for cell in row):
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_dict[h] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
    wb.close()

    def find_col(keywords):
        for h in headers:
            for kw in keywords:
                if kw.lower() in h.lower():
                    return h
        return None

    col_service = find_col(["service"]) or "Service"
    col_feature = find_col(["feature"]) or "Feature"
    col_status = find_col(["status"]) or "Status"
    col_priority = find_col(["priority"]) or "Priority"
    col_observation = find_col(["observation"]) or "Observation"
    col_recommendation = find_col(["recommendation"]) or "Recommendation"
    col_link_text = find_col(["link text", "linktext"]) or "Link Text"
    col_link_url = find_col(["link url", "linkurl"]) or "Link URL"

    # Filter non-success rows
    action_items = [r for r in rows if r.get(col_status, "").lower() and "success" not in r.get(col_status, "").lower()]
    action_items.sort(key=lambda r: priority_sort_key(r.get(col_priority, "")))

    priority_counts = {}
    for item in action_items:
        p = item.get(col_priority, "") or "Unknown"
        priority_counts[p] = priority_counts.get(p, 0) + 1

    now = datetime.now().strftime("%B %d, %Y")
    lines = []

    # ── COVER SECTION ──
    lines.append("# Microsoft 365 Copilot Readiness - Detailed Action Plan\n")
    lines.append(f"**Generated:** {now}  ")
    lines.append(f"**Total Action Items:** {len(action_items)}  ")
    lines.append(f"**Source:** M365 Copilot Readiness Assessment  ")
    lines.append(f"**Methodology:** Step 2 — Row-by-Row Implementation Guide\n")

    lines.append("## Priority Breakdown\n")
    lines.append("| Priority | Count | Target Timeline |")
    lines.append("|----------|-------|-----------------|")
    timeline = {"Critical": "Immediate (1-7 days)", "High": "30 days", "Medium": "90 days", "Low": "120 days"}
    for p in ["Critical", "High", "Medium", "Low"]:
        count = priority_counts.get(p, 0)
        if count:
            lines.append(f"| 🔴 {p} | {count} | {timeline[p]} |" if p == "Critical" else
                         f"| 🟠 {p} | {count} | {timeline[p]} |" if p == "High" else
                         f"| 🟡 {p} | {count} | {timeline[p]} |" if p == "Medium" else
                         f"| 🟢 {p} | {count} | {timeline[p]} |")
    for p, c in priority_counts.items():
        if p.lower() not in ["critical", "high", "medium", "low"]:
            lines.append(f"| ⚪ {p} | {c} | As capacity allows |")
    lines.append("")

    # ── TABLE OF CONTENTS ──
    lines.append("## Table of Contents\n")
    current_priority = None
    item_num = 0
    for item in action_items:
        item_num += 1
        feature = item.get(col_feature, "Unknown Feature")
        priority = item.get(col_priority, "Unknown")
        if priority != current_priority:
            current_priority = priority
            emoji = {"critical": "🔴", "high": "🟠", "medium": "🟡", "low": "🟢"}.get(priority.lower(), "⚪")
            lines.append(f"\n### {emoji} {priority} Priority\n")
        anchor = f"{item_num}-{feature.lower().replace(' ', '-').replace('/', '-').replace('(', '').replace(')', '')}"
        lines.append(f"- [{item_num}. {feature}](#{anchor})")
    lines.append("")
    lines.append("---\n")

    # ── DETAILED ITEMS ──
    item_num = 0
    for item in action_items:
        item_num += 1
        service = item.get(col_service, "Unknown")
        feature = item.get(col_feature, "Unknown Feature")
        priority = item.get(col_priority, "Unknown")
        status = item.get(col_status, "Unknown")
        observation = item.get(col_observation, "")
        recommendation = item.get(col_recommendation, "")
        link_text = item.get(col_link_text, "")
        link_url = item.get(col_link_url, "")

        priority_emoji = {"critical": "🔴", "high": "🟠", "medium": "🟡", "low": "🟢"}.get(priority.lower(), "⚪")

        # A. ITEM HEADER
        lines.append(f"## {item_num}. {feature}\n")
        lines.append("| Field | Value |")
        lines.append("|-------|-------|")
        lines.append(f"| **Service** | `{service}` |")
        lines.append(f"| **Priority** | {priority_emoji} {priority} |")
        lines.append(f"| **Status** | {status} |")
        lines.append("")

        # B. CURRENT SITUATION
        lines.append("### Current Situation\n")
        if observation:
            lines.append(f"> {observation}\n")
        else:
            lines.append("> _No observation details available._\n")
        # Copilot readiness context
        context = get_copilot_context(feature, status)
        lines.append(f"**Why this matters for Copilot:** {context}\n")

        # C. RECOMMENDED ACTION
        lines.append("### Recommended Action\n")
        if recommendation:
            lines.append(f"{recommendation}\n")
        else:
            lines.append("_No specific recommendation provided. Refer to Microsoft documentation for remediation guidance._\n")

        # D. RISK ASSESSMENT TABLE (generated by heuristic)
        risk = assess_risk(feature, observation, status)
        lines.append("### Risk Assessment\n")
        lines.append("| Risk Factor | Assessment |")
        lines.append("|-------------|-----------|")
        lines.append(f"| **Telemetry Limitation** | {risk['telemetry']} |")
        lines.append(f"| **False Positive Risk** | {risk['fp_risk']} |")
        lines.append(f"| **False Negative Risk** | {risk['fn_risk']} |")
        lines.append(f"| **Confidence Level** | {risk['confidence']} |")
        lines.append(f"| **Impact if Not Addressed** | {risk['impact']} |")
        lines.append("")

        # E. STEP-BY-STEP IMPLEMENTATION
        lines.append("### Step-by-Step Implementation\n")
        steps = get_steps_by_feature(feature, service)
        for i, step in enumerate(steps, 1):
            lines.append(f"{i}. {step}")
        lines.append("")

        # F. REFERENCE DOCUMENTATION
        lines.append("### Reference Documentation\n")
        if link_url and link_url.lower() not in ("none", "nan", ""):
            display = link_text if link_text and link_text.lower() not in ("none", "nan", "") else "Official Documentation"
            lines.append(f"- [{display}]({link_url})")
        else:
            lines.append("- [Microsoft Entra admin center](https://entra.microsoft.com)")
            lines.append("- [Microsoft 365 admin center](https://admin.microsoft.com)")
        lines.append("- [Microsoft Learn — M365 Security](https://learn.microsoft.com/en-us/microsoft-365/security/)")
        lines.append("")

        # G. SUCCESS CRITERIA
        lines.append("### Success Criteria\n")
        criteria = get_success_criteria(feature, service)
        for c in criteria:
            lines.append(f"- [ ] {c}")
        lines.append("")

        # H. IMPLEMENTATION NOTES
        lines.append("### Implementation Notes\n")
        lines.append("```")
        lines.append("Assigned to:             ___________________________")
        lines.append("Target completion date:   ___________________________")
        lines.append("Status:                   ⬜ Not Started | ⬜ In Progress | ⬜ Completed | ⬜ Blocked")
        lines.append("Notes:                    ")
        lines.append("```")
        lines.append("")
        lines.append("---\n")

    # ── APPENDIX ──
    lines.append("# Appendix\n")

    lines.append("## Related Resources\n")
    lines.append("| Portal | URL |")
    lines.append("|--------|-----|")
    lines.append("| Microsoft Entra admin center | https://entra.microsoft.com |")
    lines.append("| Microsoft 365 admin center | https://admin.microsoft.com |")
    lines.append("| Microsoft Defender portal | https://security.microsoft.com |")
    lines.append("| Microsoft Purview compliance | https://compliance.microsoft.com |")
    lines.append("| SharePoint admin center | https://admin.microsoft.com/sharepoint |")
    lines.append("| Microsoft Intune | https://intune.microsoft.com |")
    lines.append("| Azure portal | https://portal.azure.com |")
    lines.append("| My Security Info (users) | https://aka.ms/mysecurityinfo |")
    lines.append("| FastTrack | https://fasttrack.microsoft.com |")
    lines.append("")

    lines.append("## Priority Matrix\n")
    lines.append("| Priority | Target Timeline | Description |")
    lines.append("|----------|----------------|-------------|")
    lines.append("| 🔴 Critical | Immediate (1-7 days) | Active security gaps requiring urgent remediation |")
    lines.append("| 🟠 High | 30 days | Significant risks that should be addressed promptly |")
    lines.append("| 🟡 Medium | 90 days | Important improvements for overall readiness posture |")
    lines.append("| 🟢 Low | 120 days | Optimizations and enhancements for long-term posture |")
    lines.append("")

    lines.append("## Acronyms Glossary\n")
    lines.append("| Acronym | Full Term |")
    lines.append("|---------|-----------|")
    lines.append("| CA | Conditional Access |")
    lines.append("| MFA | Multi-Factor Authentication |")
    lines.append("| MDE | Microsoft Defender for Endpoint |")
    lines.append("| XDR | Extended Detection and Response |")
    lines.append("| DLP | Data Loss Prevention |")
    lines.append("| RBAC | Role-Based Access Control |")
    lines.append("| FIDO2 | Fast Identity Online 2 |")
    lines.append("| AIR | Automated Investigation and Response |")
    lines.append("| SPO | SharePoint Online |")
    lines.append("| OAuth | Open Authorization |")
    lines.append("| EDR | Endpoint Detection and Response |")
    lines.append("| ASR | Attack Surface Reduction |")
    lines.append("| B2B | Business-to-Business |")
    lines.append("| SCCM | System Center Configuration Manager |")
    lines.append("| SOC | Security Operations Center |")
    lines.append("| BYOD | Bring Your Own Device |")
    lines.append("")

    lines.append("## Support Contacts\n")
    lines.append("| Role | Name | Email | Phone |")
    lines.append("|------|------|-------|-------|")
    lines.append("| Project Lead | | | |")
    lines.append("| Security Architect | | | |")
    lines.append("| Identity Admin | | | |")
    lines.append("| M365 Admin | | | |")
    lines.append("| Help Desk Lead | | | |")
    lines.append("| Microsoft TAM/CSM | | | |")
    lines.append("")

    lines.append("---\n")
    lines.append(f"*Document generated on {now} | Total action items: {len(action_items)} | Methodology: M365 Copilot Assessment Insights Generator — Step 2*\n")

    # ── WRITE FILE ──
    output_path = os.path.join(output_dir, "Copilot_Readiness_Action_Plan.md")
    content = "\n".join(lines)
    with open(output_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(content)

    return output_path, len(action_items), priority_counts, len(content)


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'='*60}")
    print(f"  STEP 2: Detailed Row-by-Row Implementation Guide")
    print(f"{'='*60}")
    print(f"  Input:  {EXCEL_PATH}")
    print(f"  Output: {OUTPUT_DIR}")
    print(f"  Date:   {datetime.now().strftime('%B %d, %Y')}")
    print(f"{'='*60}\n")

    path, total, priorities, size = generate_action_plan(EXCEL_PATH, OUTPUT_DIR)

    print(f"\n{'='*60}")
    print(f"  STEP 2 COMPLETE")
    print(f"{'='*60}")
    print(f"  File: {os.path.basename(path)}")
    print(f"  Size: {size:,} bytes")
    print(f"  Total action items: {total}")
    print(f"  Priority breakdown:")
    for p in ["Critical", "High", "Medium", "Low"]:
        if priorities.get(p, 0):
            print(f"    {p}: {priorities[p]}")
    for p, c in priorities.items():
        if p.lower() not in ["critical", "high", "medium", "low"]:
            print(f"    {p}: {c}")
    print(f"\n  Sections per item:")
    print(f"    A. Item Header (service, priority, status)")
    print(f"    B. Current Situation + Copilot readiness context")
    print(f"    C. Recommended Action")
    print(f"    D. Risk Assessment (heuristic-generated)")
    print(f"    E. Step-by-Step Implementation (feature-aware)")
    print(f"    F. Reference Documentation")
    print(f"    G. Success Criteria Checklist")
    print(f"    H. Implementation Notes")
    print(f"  Appendix: Resources, Priority Matrix, Glossary, Contacts")
    print(f"{'='*60}\n")
