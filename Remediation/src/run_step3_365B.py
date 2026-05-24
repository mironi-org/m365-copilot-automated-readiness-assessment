"""
Step 3: Enterprise Project Tracker
From: M365_Copilot_Assessment_Insights_Generator.md (Option B)
Input: <input_excel> (passed as parameter)
Output: <output_directory>/Copilot_Readiness_Tracker.xlsx
"""

import os
import sys
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule

# ─── CONFIG ───────────────────────────────────────────────────────────────────
if len(sys.argv) < 3:
    print("ERROR: Both parameters are required.")
    print("  <input_excel>       - Path to the assessment Excel file (REQUIRED)")
    print("  <output_directory>  - Path to the output directory (REQUIRED)")
    print()
    print("Usage: python run_step3_365B.py <input_excel> <output_directory>")
    sys.exit(1)

EXCEL_PATH = sys.argv[1]
OUTPUT_DIR = sys.argv[2]

if not os.path.isfile(EXCEL_PATH):
    print(f"ERROR: Input file not found: {EXCEL_PATH}")
    sys.exit(1)

if not OUTPUT_DIR.strip():
    print("ERROR: Output directory cannot be empty.")
    sys.exit(1)


# ─── RISK ASSESSMENT LOGIC (aligned with reference implementation) ────────────
def assess_risk(feature, observation, status):
    """
    Generate risk assessment columns based on Status + Observation signals.
    Logic derived from original reference implementation:
    - Primary axis: Status value (how did the assessment tool collect data?)
    - Secondary axis: Observation content (what type of data was obtained?)
    Returns dict with: telemetry, fp_risk, fn_risk, confidence
    """
    f = (feature or "").lower()
    o = (observation or "").lower()
    s = (status or "").lower()

    # ── Determine FN Risk by feature family ──
    fn_risk = "Medium"

    # ── Rule 1: Access Denied — couldn't collect data at all ──
    if "access denied" in s:
        return {
            "telemetry": "Unknown",
            "fp_risk": "Medium",
            "fn_risk": "Medium",
            "confidence": "Medium",
        }

    # ── Rule 2: Warning (especially Defender threats) — partial telemetry ──
    if "warning" in s:
        return {
            "telemetry": "Telemetry Partial",
            "fp_risk": "High",
            "fn_risk": "High",
            "confidence": "Low",
        }

    # ── Rule 3: Critical status — derived posture score ──
    if "critical" in s:
        return {
            "telemetry": "Derived Posture",
            "fp_risk": "Medium",
            "fn_risk": "High",
            "confidence": "Medium",
        }

    # ── Rule 4: PendingActivation — service exists but API blocked ──
    if "pendingactivation" in s:
        return {
            "telemetry": "Permission Blocked",
            "fp_risk": "High",
            "fn_risk": "Medium",
            "confidence": "Low",
        }

    # ── Rule 5: PendingInput — needs manual verification ──
    if "pendinginput" in s:
        return {
            "telemetry": "Manual Verification Required",
            "fp_risk": "High",
            "fn_risk": "Medium",
            "confidence": "Low",
        }

    # ── Rule 6: Missing Prerequisite / Permission Required ──
    if "missing prerequisite" in s or "permission required" in s:
        return {
            "telemetry": "Unknown",
            "fp_risk": "Medium",
            "fn_risk": "Medium",
            "confidence": "Medium",
        }

    # ── Rule 7: Disabled — config absence confirmed via API ──
    if "disabled" in s:
        return {
            "telemetry": "Configuration Gap",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
        }

    # ── Rule 8: Action Required — config gap confirmed ──
    if "action required" in s:
        return {
            "telemetry": "Configuration Gap",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
        }

    # ── Rule 9: Insight — inferred from metrics/counts ──
    if "insight" in s:
        return {
            "telemetry": "Index Inferred",
            "fp_risk": "Medium",
            "fn_risk": fn_risk,
            "confidence": "Medium",
        }

    # ── Default — standard Graph API read ──
    return {
        "telemetry": "Index Inferred",
        "fp_risk": "Medium",
        "fn_risk": fn_risk,
        "confidence": "Medium",
    }


# ─── SUCCESS CRITERIA GENERATION ──────────────────────────────────────────────
def get_success_criteria_text(feature):
    f = (feature or "").lower()
    if "conditional access" in f or "ca polic" in f:
        return "CA policy created; Targets Office 365; Requires MFA; Tested 7+ days; Zero lockouts"
    elif "mfa" in f or "multi-factor" in f or "multifactor" in f:
        return "100% enrollment; Primary + backup methods; CA enforcing; Communication sent; Help desk ready"
    elif "password" in f or "passwordless" in f:
        return "Policies enabled; Pilot successful; Feedback collected; Training created; Metrics tracked"
    elif "access review" in f or "governance" in f or "review" in f:
        return "Review created; Quarterly recurrence; Reviewers assigned; First cycle done; Auto-actions set"
    elif "sharepoint" in f or "onedrive" in f or "content" in f or "site" in f:
        return "3-5 sites created; Content migrated; 100+ docs indexed; Copilot searches; Metadata applied"
    elif "defender" in f or "xdr" in f or "device" in f or "endpoint" in f:
        return "50%+ devices onboarded; Inventory visible; Policies applied; Auto investigation on; Rules created"
    elif "oauth" in f or "app" in f or "consent" in f or "application" in f:
        return "Apps reviewed; Over-privileged revoked; Consent policy set; Admin workflow on; Monitoring active"
    elif "sensitivity label" in f or "dlp" in f or "information protection" in f:
        return "Labels published; Auto-labeling configured; DLP enforcing; Simulation validated; Adoption tracked"
    elif "intune" in f or "device management" in f or "compliance polic" in f:
        return "Enrollment configured; Compliance policies assigned; CA linked; Non-compliant identified; Expanded"
    elif "cross-tenant" in f or "external" in f or "b2b" in f or "guest" in f:
        return "Settings reviewed; MFA trust enabled; CA for guests active; B2B settings aligned; Monitoring on"
    elif "group" in f or "licens" in f:
        return "Groups created; Licenses assigned via groups; Errors resolved; Direct assignments removed; Monitored"
    else:
        return "Config completed; Validated; No adverse impact; Docs updated; Next assessment improved"


# ─── STYLES ───────────────────────────────────────────────────────────────────
NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ALT_ROW_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)

PRIORITY_FILLS = {
    "Critical": (PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"), Font(name="Calibri", size=11, bold=True, color="FFFFFF")),
    "High": (PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"), Font(name="Calibri", size=11, bold=True, color="FFFFFF")),
    "Medium": (PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"), Font(name="Calibri", size=11, bold=True, color="000000")),
    "Low": (PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), Font(name="Calibri", size=11, bold=True, color="000000")),
}

STATUS_FILLS = {
    "Not Started": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "In Progress": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Completed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Blocked": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

WRAP_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")


# ─── COLUMN DEFINITIONS ──────────────────────────────────────────────────────
COLUMNS = [
    ("Item #", 8),
    ("Priority", 10),
    ("Status", 18),
    ("Service", 12),
    ("Feature", 35),
    ("Observation", 50),
    ("Recommendation", 50),
    ("Telemetry Limitation Detected", 15),
    ("False Positive Risk", 15),
    ("False Negative Risk", 18),
    ("Confidence in Observation", 18),
    ("Link Text", 25),
    ("Link URL", 45),
    ("Success Criteria Checklist", 40),
    ("Assigned To", 18),
    ("Target Completion Date", 18),
    ("Actual Completion Date", 18),
    ("Implementation Status", 18),
    ("Progress %", 12),
    ("Implementation Notes", 40),
    ("Blockers/Issues", 30),
]


def priority_sort_key(p):
    return {"critical": 0, "high": 1, "medium": 2, "low": 3}.get((p or "").lower(), 99)


# ─── HELPER: APPLY FORMATTING TO A TRACKER SHEET ─────────────────────────────
def format_tracker_sheet(ws, data_rows, add_validation=True):
    """Apply headers, widths, formatting, filters, validation to a tracker sheet."""
    # Write header row
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

            # Alternating row fill
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

        # Priority formatting (col 2)
        priority_val = row_data[1] if len(row_data) > 1 else ""
        if priority_val in PRIORITY_FILLS:
            fill, font = PRIORITY_FILLS[priority_val]
            cell_p = ws.cell(row=row_idx, column=2)
            cell_p.fill = fill
            cell_p.font = font

        # Wrap text for long columns
        for wrap_col in [6, 7, 14, 20, 21]:
            if wrap_col <= len(row_data):
                ws.cell(row=row_idx, column=wrap_col).alignment = WRAP_ALIGN

        # Link URL formatting (col 13)
        if len(row_data) >= 13 and row_data[12]:
            link_cell = ws.cell(row=row_idx, column=13)
            link_cell.font = LINK_FONT
            url = row_data[12]
            if url and url.startswith("http"):
                link_cell.hyperlink = url

    # Freeze panes
    ws.freeze_panes = "C2"

    # Auto-filter
    if data_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(data_rows) + 1}"

    # Data validation for Implementation Status (col 18)
    if add_validation and data_rows:
        status_dv = DataValidation(
            type="list",
            formula1='"Not Started,In Progress,Completed,Blocked"',
            allow_blank=True,
        )
        status_dv.error = "Please select a valid status"
        status_dv.errorTitle = "Invalid Status"
        status_dv.showDropDown = False
        ws.add_data_validation(status_dv)
        status_dv.add(f"R2:R{len(data_rows) + 1}")

        # Progress % validation (col 19)
        progress_dv = DataValidation(
            type="whole",
            operator="between",
            formula1="0",
            formula2="100",
            allow_blank=True,
        )
        progress_dv.error = "Please enter a value between 0 and 100"
        progress_dv.errorTitle = "Invalid Progress"
        ws.add_data_validation(progress_dv)
        progress_dv.add(f"S2:S{len(data_rows) + 1}")

    # Conditional formatting — skip if no data rows (range would be invalid)
    if len(data_rows) == 0:
        return

    # Conditional formatting for Priority column
    ws.conditional_formatting.add(
        f"B2:B{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"), font=Font(bold=True, color="FFFFFF")),
    )
    ws.conditional_formatting.add(
        f"B2:B{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"), font=Font(bold=True, color="FFFFFF")),
    )
    ws.conditional_formatting.add(
        f"B2:B{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"), font=Font(bold=True, color="000000")),
    )
    ws.conditional_formatting.add(
        f"B2:B{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), font=Font(bold=True, color="000000")),
    )

    # Conditional formatting for Implementation Status (col 18)
    ws.conditional_formatting.add(
        f"R2:R{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Not Started"'], fill=PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        f"R2:R{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), font=Font(bold=True)),
    )
    ws.conditional_formatting.add(
        f"R2:R{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Completed"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), font=Font(bold=True)),
    )
    ws.conditional_formatting.add(
        f"R2:R{len(data_rows) + 1}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(bold=True)),
    )


# ─── MAIN GENERATION ─────────────────────────────────────────────────────────
def generate_tracker(excel_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    # Read source data
    wb_src = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws_src = wb_src.active
    headers = [str(cell.value or "").strip() for cell in ws_src[1]]

    rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row and any(cell is not None for cell in row):
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_dict[h] = str(val).strip() if val is not None else ""
            rows.append(row_dict)
    wb_src.close()

    def find_col(keywords):
        for h in headers:
            for kw in keywords:
                if kw.lower() in h.lower():
                    return h
        return None

    col_service = find_col(["service"]) or "Service"
    col_feature = find_col(["feature"]) or "Feature"
    col_status = find_col(["status"]) or "Status"
    col_priority = find_col(["priority"]) or "Priority"
    col_observation = find_col(["observation"]) or "Observation"
    col_recommendation = find_col(["recommendation"]) or "Recommendation"
    col_link_text = find_col(["link text", "linktext"]) or "Link Text"
    col_link_url = find_col(["link url", "linkurl"]) or "Link URL"

    # Filter non-success rows
    action_items = [r for r in rows if r.get(col_status, "").lower() and "success" not in r.get(col_status, "").lower()]
    action_items.sort(key=lambda r: priority_sort_key(r.get(col_priority, "")))

    # Build tracker data rows (21 columns)
    tracker_data = []
    for idx, item in enumerate(action_items, 1):
        feature = item.get(col_feature, "")
        observation = item.get(col_observation, "")
        status = item.get(col_status, "")
        risk = assess_risk(feature, observation, status)

        row = [
            idx,                                        # 1 Item #
            item.get(col_priority, ""),                 # 2 Priority
            status,                                     # 3 Status
            item.get(col_service, ""),                  # 4 Service
            feature,                                    # 5 Feature
            observation,                                # 6 Observation
            item.get(col_recommendation, ""),           # 7 Recommendation
            risk["telemetry"],                          # 8 Telemetry Limitation
            risk["fp_risk"],                            # 9 False Positive Risk
            risk["fn_risk"],                            # 10 False Negative Risk
            risk["confidence"],                         # 11 Confidence
            item.get(col_link_text, ""),                # 12 Link Text
            item.get(col_link_url, ""),                 # 13 Link URL
            get_success_criteria_text(feature),         # 14 Success Criteria
            "",                                         # 15 Assigned To
            "",                                         # 16 Target Completion Date
            "",                                         # 17 Actual Completion Date
            "Not Started",                              # 18 Implementation Status
            0,                                          # 19 Progress %
            "",                                         # 20 Implementation Notes
            "",                                         # 21 Blockers/Issues
        ]
        tracker_data.append(row)

    # Create workbook
    wb = openpyxl.Workbook()

    # ─── SHEET 1: All Items Tracker ───
    ws1 = wb.active
    ws1.title = "All Items Tracker"
    format_tracker_sheet(ws1, tracker_data)

    # ─── SHEET 2: Critical & High Priority ───
    ws2 = wb.create_sheet("Critical & High Priority")
    crit_high = [r for r in tracker_data if (r[1] or "").lower() in ("critical", "high")]
    # Re-number
    crit_high_renumbered = []
    for i, row in enumerate(crit_high, 1):
        new_row = list(row)
        new_row[0] = i
        crit_high_renumbered.append(new_row)
    format_tracker_sheet(ws2, crit_high_renumbered)

    # ─── SHEET 3: Summary Dashboard ───
    ws3 = wb.create_sheet("Summary Dashboard")
    _build_dashboard(ws3, action_items, tracker_data, col_priority, col_status, col_service)

    # ─── SHEET 4: By Service ───
    ws4 = wb.create_sheet("By Service")
    _build_by_service(ws4, action_items, col_service, col_priority)

    # ─── SHEET 5: Instructions ───
    ws5 = wb.create_sheet("Instructions")
    _build_instructions(ws5, len(tracker_data))

    # Save
    output_path = os.path.join(output_dir, "Copilot_Readiness_Tracker.xlsx")
    wb.save(output_path)
    return output_path, len(tracker_data), len(crit_high)


# ─── SHEET 3: DASHBOARD ──────────────────────────────────────────────────────
def _build_dashboard(ws, items, tracker_data, col_priority, col_status, col_service):
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    metric_font = Font(name="Calibri", size=11, bold=True)
    value_font = Font(name="Calibri", size=11)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 12

    row = 1
    ws.cell(row=row, column=2, value="Microsoft 365 Copilot Readiness - Summary Dashboard").font = title_font
    ws.merge_cells("B1:F1")
    row += 1
    ws.cell(row=row, column=2, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}").font = Font(italic=True, size=10)
    row += 2

    # Priority counts
    priorities = {}
    for item in items:
        p = item.get(col_priority, "") or "Unknown"
        priorities[p] = priorities.get(p, 0) + 1

    # Status counts
    statuses = {}
    for item in items:
        s = item.get(col_status, "") or "Unknown"
        statuses[s] = statuses.get(s, 0) + 1

    # Service counts
    services = {}
    for item in items:
        svc = item.get(col_service, "") or "Unknown"
        services[svc] = services.get(svc, 0) + 1

    # Metrics section
    ws.cell(row=row, column=2, value="Key Metrics").font = section_font
    row += 1

    metrics = [
        ("Total Items", len(items)),
        ("Critical Priority", priorities.get("Critical", 0)),
        ("High Priority", priorities.get("High", 0)),
        ("Medium Priority", priorities.get("Medium", 0)),
        ("Low Priority", priorities.get("Low", 0)),
    ]
    for label, val in metrics:
        ws.cell(row=row, column=2, value=label).font = metric_font
        ws.cell(row=row, column=3, value=val).font = value_font
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Status Breakdown").font = section_font
    row += 1
    for label, val in sorted(statuses.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=2, value=label).font = metric_font
        ws.cell(row=row, column=3, value=val).font = value_font
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="Service Breakdown").font = section_font
    row += 1
    for label, val in sorted(services.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=2, value=label).font = metric_font
        ws.cell(row=row, column=3, value=val).font = value_font
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        row += 1

    # ─── CHARTS ───
    # Priority Distribution Pie Chart
    chart_start_row = row + 2
    ws.cell(row=chart_start_row, column=2, value="Priority").font = Font(bold=True)
    ws.cell(row=chart_start_row, column=3, value="Count").font = Font(bold=True)
    chart_row = chart_start_row + 1
    for p in ["Critical", "High", "Medium", "Low"]:
        if priorities.get(p, 0):
            ws.cell(row=chart_row, column=2, value=p)
            ws.cell(row=chart_row, column=3, value=priorities[p])
            chart_row += 1

    if chart_row > chart_start_row + 1:
        pie = PieChart()
        pie.title = "Priority Distribution"
        pie.style = 10
        data_ref = Reference(ws, min_col=3, min_row=chart_start_row, max_row=chart_row - 1)
        cats_ref = Reference(ws, min_col=2, min_row=chart_start_row + 1, max_row=chart_row - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.width = 14
        pie.height = 10
        ws.add_chart(pie, "E4")

    # Status Distribution Bar Chart
    bar_start = chart_row + 2
    ws.cell(row=bar_start, column=2, value="Status").font = Font(bold=True)
    ws.cell(row=bar_start, column=3, value="Count").font = Font(bold=True)
    bar_row = bar_start + 1
    for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
        ws.cell(row=bar_row, column=2, value=s)
        ws.cell(row=bar_row, column=3, value=c)
        bar_row += 1

    if bar_row > bar_start + 1:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Status Distribution"
        bar.style = 10
        data_ref = Reference(ws, min_col=3, min_row=bar_start, max_row=bar_row - 1)
        cats_ref = Reference(ws, min_col=2, min_row=bar_start + 1, max_row=bar_row - 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 14
        bar.height = 10
        ws.add_chart(bar, "E20")


# ─── SHEET 4: BY SERVICE ─────────────────────────────────────────────────────
def _build_by_service(ws, items, col_service, col_priority):
    # Headers
    headers_s = ["Service", "Priority", "Count"]
    for c, h in enumerate(headers_s, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    # Group by service, priority
    grouped = {}
    for item in items:
        svc = item.get(col_service, "") or "Unknown"
        pri = item.get(col_priority, "") or "Unknown"
        key = (svc, pri)
        grouped[key] = grouped.get(key, 0) + 1

    # Sort: by service name, then priority
    sorted_groups = sorted(grouped.items(), key=lambda x: (x[0][0], priority_sort_key(x[0][1])))

    row = 2
    for (svc, pri), count in sorted_groups:
        ws.cell(row=row, column=1, value=svc).border = THIN_BORDER
        cell_p = ws.cell(row=row, column=2, value=pri)
        cell_p.border = THIN_BORDER
        cell_p.alignment = CENTER_ALIGN
        if pri in PRIORITY_FILLS:
            fill, font = PRIORITY_FILLS[pri]
            cell_p.fill = fill
            cell_p.font = font
        ws.cell(row=row, column=3, value=count).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        if row % 2 == 0:
            for c in range(1, 4):
                if ws.cell(row=row, column=c).fill == PatternFill():
                    ws.cell(row=row, column=c).fill = ALT_ROW_FILL
        row += 1

    # Auto-filter
    if row > 2:
        ws.auto_filter.ref = f"A1:C{row - 1}"

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


# ─── SHEET 5: INSTRUCTIONS ───────────────────────────────────────────────────
def _build_instructions(ws, total_items):
    """Two-column user guide matching reference formatting."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100

    # Fonts
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    section_label_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=10, bold=True)
    body_font = Font(name="Calibri", size=11)

    # Fills
    section_fill = PatternFill("solid", fgColor="1F4E78")
    label_fill = PatternFill("solid", fgColor="E7E6E6")

    wrap_align = Alignment(vertical="center", wrap_text=True)

    # Row data: (col_a_text, col_b_text, is_section_header)
    rows = [
        ("Copilot Readiness Tracker - User Guide", None, True),  # title row (merged)
        ("Overview",
         f"This tracker contains all {total_items} observations from the M365 Copilot Readiness Assessment. "
         "Use it to plan, assign, track, and validate remediation efforts.",
         "section"),
        ("How to Use This Tracker",
         '1. Review all items in "All Items Tracker" sheet\n'
         '2. Focus on "Critical & High Priority" sheet for immediate actions\n'
         '3. Assign owners using "Assigned To" column\n'
         "4. Set target dates based on priority timelines\n"
         '5. Update "Implementation Status" and "Progress %" regularly\n'
         '6. Use "Success Criteria Checklist" to validate completion\n'
         '7. Review "Summary Dashboard" for overall progress',
         False),
        ("Column Descriptions - Item #",
         f"Unique identifier for each remediation item (1-{total_items})", False),
        ("Column Descriptions - Priority",
         "Critical (immediate), High (7-30 days), Medium (30-90 days), Low (90+ days)", False),
        ("Column Descriptions - Status",
         "Current state: Action Required, Insight, Warning, PendingActivation, Permission Required, etc.", False),
        ("Column Descriptions - Implementation Status",
         "Your tracking status: Not Started, In Progress, Completed, Blocked (use dropdown)", False),
        ("Column Descriptions - Progress %",
         "Percentage complete (0-100%). Update as you progress through implementation steps.", False),
        ("Column Descriptions - Success Criteria",
         "Key validation points to confirm successful implementation. Check each criterion before marking as complete.", False),
        ("Tracking Process",
         "1. Weekly team review of tracker\n"
         "2. Update status and progress after each work session\n"
         "3. Document blockers immediately\n"
         "4. Escalate blocked items to stakeholders\n"
         "5. Validate with success criteria before marking complete\n"
         "6. Schedule follow-up assessment to verify improvements", False),
        ("Priority Timeline",
         "Critical: 0-7 days - Immediate security risks blocking Copilot deployment", False),
        ("Priority Timeline",
         "High: 7-30 days - Essential security and identity configurations", False),
        ("Priority Timeline",
         "Medium: 30-90 days - Optimization and governance improvements", False),
        ("Priority Timeline",
         "Low: 90+ days - Long-term enhancements and best practices", False),
        ("Status Definitions",
         "Action Required = Immediate configuration needed\n"
         "Insight = Optimization opportunity\n"
         "Warning = Potential risk detected\n"
         "PendingActivation = License upgrade required\n"
         "Permission Required = API access needed", False),
        ("Tips for Success",
         "\u2022 Start with Critical & High priority items\n"
         "\u2022 Assign realistic timelines\n"
         "\u2022 Test changes in pilot before broad rollout\n"
         "\u2022 Document all configuration changes\n"
         "\u2022 Communicate changes to affected users\n"
         "\u2022 Keep stakeholders informed of progress", False),
        ("Support Resources",
         "Markdown Guide: Copilot_Readiness_Action_Plan.md\n"
         "Microsoft Learn: https://learn.microsoft.com/microsoft-365-copilot/\n"
         "Entra Admin: https://entra.microsoft.com\n"
         "M365 Admin: https://admin.microsoft.com\n"
         "Security Portal: https://security.microsoft.com", False),
    ]

    for row_idx, (col_a, col_b, row_type) in enumerate(rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=col_a)
        cell_a.alignment = wrap_align

        if row_type is True:
            # Title row — merged, large font
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell_a.font = title_font
            ws.row_dimensions[row_idx].height = 25
        elif row_type == "section":
            # Section header row — white on blue for both columns
            cell_a.font = section_label_font
            cell_a.fill = section_fill
            cell_b = ws.cell(row=row_idx, column=2, value=col_b)
            cell_b.font = section_label_font
            cell_b.fill = section_fill
            cell_b.alignment = wrap_align
            ws.row_dimensions[row_idx].height = 60
        else:
            # Normal data row — gray label, white body
            cell_a.font = label_font
            cell_a.fill = label_fill
            cell_b = ws.cell(row=row_idx, column=2, value=col_b)
            cell_b.font = body_font
            cell_b.alignment = wrap_align
            ws.row_dimensions[row_idx].height = 60


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{'='*60}")
    print(f"  STEP 3: Enterprise Project Tracker")
    print(f"{'='*60}")
    print(f"  Input:  {EXCEL_PATH}")
    print(f"  Output: {OUTPUT_DIR}")
    print(f"  Date:   {datetime.now().strftime('%B %d, %Y')}")
    print(f"{'='*60}\n")

    path, total, crit_high = generate_tracker(EXCEL_PATH, OUTPUT_DIR)

    size = os.path.getsize(path)
    print(f"\n{'='*60}")
    print(f"  STEP 3 COMPLETE")
    print(f"{'='*60}")
    print(f"  File: {os.path.basename(path)}")
    print(f"  Size: {size:,} bytes")
    print(f"  Total action items: {total}")
    print(f"  Critical + High priority items: {crit_high}")
    print(f"\n  Workbook structure (5 sheets):")
    print(f"    1. All Items Tracker — {total} rows, 21 columns")
    print(f"    2. Critical & High Priority — {crit_high} rows (filtered)")
    print(f"    3. Summary Dashboard — Metrics + Pie/Bar charts")
    print(f"    4. By Service — Grouped counts for team assignment")
    print(f"    5. Instructions — Full user guide")
    print(f"\n  Features:")
    print(f"    ✓ Navy headers, frozen panes, auto-filters")
    print(f"    ✓ Conditional formatting (Priority + Status colors)")
    print(f"    ✓ Data validation dropdowns (Implementation Status)")
    print(f"    ✓ Progress % validation (0-100)")
    print(f"    ✓ AI-generated risk assessment columns (8-11)")
    print(f"    ✓ Feature-specific success criteria (col 14)")
    print(f"    ✓ Hyperlinks in Link URL column")
    print(f"    ✓ Alternating row shading")
    print(f"{'='*60}\n")
